import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

load_wb = load_workbook(r'이메일.xlsx', data_only=True)
load_ws = load_wb.active

for i in range(1, load_ws.max_row+1) :
    recv_email_value=load_ws.cell(i,1).value
    print("성공 : ",recv_email_value)
    
    try :
        send_email = "구글 아이디@naver.com"
        send_pwd = "구글앱 비밀번호"

        recv_email = "받는 이메일@hanmail.net"

        smtp_name = "smtp.naver.com"
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] = "엑셀 파일에서 읽어온 이메일에 보내집니다."
        msg['From'] = send_email
        msg['To'] = recv_email

        text = """
        메일 내용
        """

        msg.attach(MIMEText(text))

        s = smtplib.SMTP(smtp_name, smtp_port)
        s.starttls()
        s.login(send_email, send_pwd)
        s.sendmail(send_email, recv_email, msg.as_string())
        s.quit()
    except :
        print("에러 : ", recv_email_value) # 에러가 발생한 메일 주소 출력