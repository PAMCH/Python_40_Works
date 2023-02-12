from openpyxl import load_workbook # 엑셀 파일을 불러오기 위한 모듈

load_wb = load_workbook(r'수료증명단.xlsx') # 엑셀 파일을 읽어봐서 할당
load_ws = load_wb.active # 읽어온 엑셀 파일의 활성화된 부분 할당

name_list = []
birthday_list = []
ho_list = []

for i in range(1, load_ws.max_row + 1) :
    name_list.append(load_ws.cell(i,1).value)
    birthday_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)

print(name_list)
print(birthday_list)
print(ho_list)