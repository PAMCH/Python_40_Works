import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://v.daum.net/v/20230213200007930' # 스크랩을 위한 뉴스기사 url

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url,headers=headers)

res = re.findall(r'[\w\.-]+@[\w\.-]',response.text)

resL = list(set(res))

print(res[0])

try :
    wb = load_workbook(r'\\email.xlsx', data_only=True) # 엑셀 파일이 있다면 읽어옴
    sheet = wb.active
except :
    wb =Workbook() # 엑셀 파일이 없으면 새로 생성
    sheet = wb.active

for res in resL : # 수집된 메일 수 만큼 반복함
    sheet.append([res])

wb.save(r'\\email.xlsx')